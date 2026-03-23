"""PDF to Excel Converter Tool — PySide6.

Two-panel layout: left settings panel, right page preview + report.
Bottom: scrollable thumbnail strip.
"""

from __future__ import annotations

import logging
import os
import subprocess
import sys
from pathlib import Path
from typing import Optional

from PySide6.QtWidgets import (
    QWidget,
    QFrame,
    QLabel,
    QPushButton,
    QLineEdit,
    QComboBox,
    QCheckBox,
    QSpinBox,
    QScrollArea,
    QHBoxLayout,
    QVBoxLayout,
    QStackedWidget,
    QFileDialog,
    QMessageBox,
    QInputDialog,
    QProgressBar,
    QSizePolicy,
)
from PySide6.QtCore import Qt, QThread, QTimer, QSize, Signal
from PySide6.QtGui import QPainter, QColor, QPixmap, QPen, QFont, QIcon
from icons import svg_pixmap
from colors import (
    WHITE,
    G100,
    G200,
    G300,
    G400,
    G500,
    G700,
    G900,
    BLUE,
    BLUE_HOVER,
    GREEN,
    GREEN_HOVER,
    RED,
    THUMB_BG,
    EMERALD,
    BLUE_MED,)
from utils import _fitz_pix_to_qpixmap, _WheelToHScroll, assert_file_writable

try:
    import fitz

    _HAS_FITZ = True
except ImportError:
    _HAS_FITZ = False

try:
    import pdfplumber

    _HAS_PLUMBER = True
except ImportError:
    _HAS_PLUMBER = False

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    openpyxl = None

logger = logging.getLogger(__name__)


class _ExcelExtractionWorker(QThread):
    progress = Signal(int, str)  # percent, status text
    finished = Signal(str, str)  # report_text, out_dir
    failed = Signal(str)

    def __init__(
        self,
        pdf_path,
        password,
        pages,
        out_path,
        settings,
        min_rows,
        min_cols,
        skip_image,
        sheet_mode,
        bold_header,
        auto_fit,
        parent=None,
    ):
        super().__init__(parent)
        self._pdf_path = pdf_path
        self._password = password
        self._pages = pages
        self._out_path = out_path
        self._settings = settings
        self._min_rows = min_rows
        self._min_cols = min_cols
        self._skip_image = skip_image
        self._sheet_mode = sheet_mode
        self._bold_header = bold_header
        self._auto_fit = auto_fit

    @staticmethod
    def _page_is_image_only(doc, pg_idx):
        raw = doc[pg_idx].get_text()
        text = raw.strip() if isinstance(raw, str) else ""
        return len(text) < 5

    def run(self):
        try:
            assert_file_writable(Path(self._out_path))
            self._do_extract()
        except PermissionError as exc:
            self.failed.emit(str(exc))
        except Exception as exc:
            self.failed.emit(str(exc))

    def _do_extract(self):
        doc = fitz.open(self._pdf_path)
        pldoc = pdfplumber.open(self._pdf_path)
        if self._password:
            doc.authenticate(self._password)

        try:
            warnings = []
            report_lines = ["=== Extraction Complete ===\n"]
            report_lines.append(f"Input:  {os.path.basename(self._pdf_path)}")
            report_lines.append(
                f"Pages processed: {len(self._pages)}  "
                f"(pages {self._pages[0] + 1}\u2013{self._pages[-1] + 1})\n"
            )

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            total_tables = 0
            total_rows = 0
            sheet_idx = 0
            estimated_total = max(len(self._pages), 1)

            for pg_num, pg_idx in enumerate(self._pages):
                self.progress.emit(
                    int(pg_num / estimated_total * 85),
                    f"Processing page {pg_idx + 1}…",
                )

                if self._skip_image and self._page_is_image_only(doc, pg_idx):
                    warnings.append(f"Page {pg_idx + 1}: image-only page — skipped.")
                    continue

                try:
                    pl_page = pldoc.pages[pg_idx]
                    tables = pl_page.find_tables(self._settings)
                except Exception as e:
                    warnings.append(f"Page {pg_idx + 1}: detection error — {e}")
                    continue

                tables = [
                    t
                    for t in tables
                    if t.rows
                    and len(t.rows) >= self._min_rows
                    and len(t.rows[0]) >= self._min_cols
                ]

                if not tables:
                    warnings.append(f"Page {pg_idx + 1}: no qualifying tables found.")

                for tbl_idx, tbl in enumerate(tables):
                    rows = tbl.extract()
                    if not rows:
                        continue

                    if self._sheet_mode == "Table":
                        sheet_name = f"P{pg_idx + 1}_T{tbl_idx + 1}"
                    else:
                        sheet_name = f"Page {pg_idx + 1}"

                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=sheet_name)
                        first_row_written = False
                    else:
                        ws = wb[sheet_name]
                        first_row_written = True

                    for row_idx, row in enumerate(rows):
                        cells = [str(c) if c is not None else "" for c in row]
                        ws.append(cells)
                        if row_idx == 0 and not first_row_written and self._bold_header:
                            for cell in ws[ws.max_row]:
                                cell.font = XLFont(bold=True)
                                cell.fill = PatternFill("solid", fgColor="DBEAFE")

                    if self._auto_fit:
                        for col in ws.columns:
                            max_len = max(
                                (len(str(cell.value or "")) for cell in col), default=8
                            )
                            ws.column_dimensions[col[0].column_letter].width = min(
                                max_len + 2, 50
                            )

                    total_tables += 1
                    total_rows += len(rows)
                    sheet_idx += 1
                    report_lines.append(
                        f"Table {total_tables} — page {pg_idx + 1}, sheet '{sheet_name}'"
                    )
                    report_lines.append(
                        f"  Dimensions: {len(rows)} rows \u00d7 "
                        f"{max(len(r) for r in rows) if rows else 0} columns\n"
                    )

            if not wb.sheetnames:
                wb.create_sheet("No tables found")
                warnings.append("No tables were found in the selected pages.")

            wb.save(self._out_path)

            report_lines.append(
                "\n\u2500\u2500 Summary \u2500\u2500\u2500\u2500\u2500\u2500\u2500"
                "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
                "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
            )
            report_lines.append(f"Tables extracted: {total_tables}")
            report_lines.append(f"Total rows:       {total_rows}")
            report_lines.append(f"Output file:      {os.path.basename(self._out_path)}")

            if warnings:
                report_lines.append(
                    "\n\u2500\u2500 Warnings \u2500\u2500\u2500\u2500\u2500\u2500\u2500"
                    "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
                    "\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500"
                )
                for w in warnings:
                    report_lines.append(f"  \u2022 {w}")

            out_dir = str(Path(self._out_path).parent)
            self.finished.emit("\n".join(report_lines), out_dir)

        finally:
            try:
                doc.close()
            except Exception:
                pass
            try:
                pldoc.close()
            except Exception:
                pass


def _render_page_qpixmap(doc, idx: int, max_w: int):
    page = doc[idx]
    scale = max_w / page.rect.width
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    return _fitz_pix_to_qpixmap(pix), scale


def _render_thumb_qpixmap(doc, idx: int, thumb_w: int) -> QPixmap:
    page = doc[idx]
    scale = thumb_w / page.rect.width
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    return _fitz_pix_to_qpixmap(pix)


class _PreviewCanvas(QWidget):
    def __init__(self, tool: "PDFToExcelTool", parent=None):
        super().__init__(parent)
        self._t = tool
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.setMinimumSize(300, 300)

    def paintEvent(self, _):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        t = self._t
        p.fillRect(self.rect(), QColor(THUMB_BG))
        if t._page_pixmap is None:
            p.setPen(QColor(G400))
            p.setFont(QFont("Segoe UI", 13))
            p.drawText(
                self.rect(),
                Qt.AlignmentFlag.AlignCenter,
                "Open a PDF to preview it here",
            )
            return
        p.drawPixmap(t._page_ox, t._page_oy, t._page_pixmap)
        pen = QPen(QColor(BLUE), 2, Qt.PenStyle.DashLine)
        p.setPen(pen)
        p.setBrush(Qt.BrushStyle.NoBrush)
        for cx0, cy0, cx1, cy1 in t._canvas_table_rects:
            p.drawRect(int(cx0), int(cy0), int(cx1 - cx0), int(cy1 - cy0))

    def resizeEvent(self, e):
        super().resizeEvent(e)
        if self._t._doc:
            self._t._render_page_canvas()


class PDFToExcelTool(QWidget):
    LEFT_W = 360
    THUMB_W = 80

    def __init__(self, parent=None):
        super().__init__(parent)

        self.pdf_path: str = ""
        self._password: str = ""
        self._doc = None
        self._pldoc = None
        self._total_pages = 0
        self._current_page = 0
        self._thumb_pixmaps: list = []
        self._thumb_render_next = 0
        self._thumb_timer: Optional[QTimer] = None
        self._highlighted_thumb_frame = None
        self._page_pixmap: Optional[QPixmap] = None
        self._page_scale = 1.0
        self._page_ox = 0
        self._page_oy = 0
        self._table_bboxes: list = []
        self._canvas_table_rects: list = []
        self._report_widget: Optional[QWidget] = None
        self._worker = None

        root_v = QVBoxLayout(self)
        root_v.setContentsMargins(0, 0, 0, 0)
        root_v.setSpacing(0)

        top = QWidget()
        top.setStyleSheet("background: transparent;")
        top_h = QHBoxLayout(top)
        top_h.setContentsMargins(0, 0, 0, 0)
        top_h.setSpacing(0)
        root_v.addWidget(top, 1)

        self._strip_container = QWidget()
        self._strip_container.setFixedHeight(155)
        self._strip_container.setStyleSheet(f"background: {G100};")
        root_v.addWidget(self._strip_container)

        missing = []
        if not _HAS_FITZ:
            missing.append("pymupdf  (pip install pymupdf)")
        if not _HAS_PLUMBER:
            missing.append("pdfplumber  (pip install pdfplumber)")
        if not _HAS_OPENPYXL:
            missing.append("openpyxl  (pip install openpyxl)")

        if missing:
            self._build_missing_deps(top, missing)
        else:
            self._build_left(top_h)
            self._build_right(top_h)
            self._build_thumb_strip()

    def _build_missing_deps(self, top: QWidget, missing: list):
        w = QWidget()
        w.setStyleSheet(f"background: {WHITE};")
        lay = QVBoxLayout(w)
        msg = "Missing dependencies:\n" + "\n".join(f"  \u2022 {m}" for m in missing)
        lbl = QLabel(msg)
        lbl.setStyleSheet(f"color: {RED}; font: 14px 'Segoe UI';")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl)
        top_lay = top.layout()
        if top_lay is not None:
            top_lay.addWidget(w)

    def _build_left(self, parent_h: QHBoxLayout):
        left_outer = QWidget()
        left_outer.setFixedWidth(self.LEFT_W)
        left_outer.setStyleSheet(
            f"background: {WHITE}; border-right: 1px solid {G200};"
        )
        outer_v = QVBoxLayout(left_outer)
        outer_v.setContentsMargins(0, 0, 0, 0)
        outer_v.setSpacing(0)
        parent_h.addWidget(left_outer)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"border: none; background: {WHITE};")
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        inner = QWidget()
        inner.setStyleSheet(f"background: {WHITE};")
        scroll.setWidget(inner)
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(24, 24, 24, 12)
        lay.setSpacing(0)

        # Title
        title_row = QHBoxLayout()
        title_row.setSpacing(12)
        title_row.setContentsMargins(0, 0, 0, 0)

        icon_box = QLabel()
        icon_box.setFixedSize(40, 40)
        icon_box.setPixmap(svg_pixmap("file-text", BLUE, 20))
        icon_box.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_box.setStyleSheet(f"background: {BLUE_MED}; border-radius: 8px;")
        title_row.addWidget(icon_box)

        title_lbl = QLabel("PDF to Excel")
        title_lbl.setStyleSheet(
            f"color: {G900}; font: bold 20px; background: transparent; border: none;"
        )
        title_row.addWidget(title_lbl)
        title_row.addStretch()
        lay.addLayout(title_row)
        lay.addSpacing(24)

        # Source file
        self._section(lay, "SOURCE FILE")
        lay.addSpacing(8)

        drop_zone = QFrame()
        drop_zone.setFixedHeight(56)
        drop_zone.setStyleSheet(
            f"background: {G100};"
            f" border: 2px dashed {G200}; border-radius: 12px;"
        )
        dz_lay = QHBoxLayout(drop_zone)
        dz_lay.setContentsMargins(10, 0, 10, 0)
        dz_lay.setSpacing(8)

        dz_icon = QLabel()
        dz_icon.setPixmap(svg_pixmap("file-text", G400, 20))
        dz_icon.setStyleSheet("border: none; background: transparent;")
        dz_lay.addWidget(dz_icon)

        self._file_entry = QLineEdit()
        self._file_entry.setReadOnly(True)
        self._file_entry.setPlaceholderText("Drop PDF here or browse…")
        self._file_entry.setStyleSheet(
            f"border: none; background: transparent; color: {G500}; font: 13px;"
        )
        dz_lay.addWidget(self._file_entry, 1)

        browse_btn = QPushButton("Browse")
        browse_btn.setFixedSize(72, 32)
        browse_btn.setStyleSheet(
            f"QPushButton {{background: {BLUE}; color: white; border-radius: 6px;"
            f" font: 13px;}} QPushButton:hover {{background: {BLUE_HOVER};}}"
        )
        browse_btn.clicked.connect(self._browse_file)
        dz_lay.addWidget(browse_btn)
        lay.addWidget(drop_zone)
        lay.addSpacing(20)

        # Detection method
        self._section(lay, "DETECTION METHOD")
        lay.addSpacing(6)
        self._detect_combo = QComboBox()
        self._detect_combo.addItems(
            ["Lattice (ruled lines)", "Stream (whitespace)", "Hybrid"]
        )
        self._detect_combo.setFixedHeight(30)
        self._detect_combo.setStyleSheet(self._combo_style())
        lay.addWidget(self._detect_combo)
        lay.addSpacing(16)

        # Page range
        self._section(lay, "PAGE RANGE")
        lay.addSpacing(6)
        self._range_entry = QLineEdit()
        self._range_entry.setPlaceholderText("All pages (e.g. 1-3, 5)")
        self._range_entry.setFixedHeight(30)
        self._range_entry.setStyleSheet(
            f"border: 1px solid {G200}; border-radius: 6px; padding: 0 8px;"
            f" font: 13px; color: {G900}; background: {WHITE};"
        )
        lay.addWidget(self._range_entry)
        lay.addSpacing(16)

        # Filters
        self._section(lay, "FILTERS")
        lay.addSpacing(8)

        min_row_w = QWidget()
        min_row_w.setStyleSheet("background: transparent;")
        mr_h = QHBoxLayout(min_row_w)
        mr_h.setContentsMargins(0, 0, 0, 0)
        mr_h.setSpacing(8)
        mr_lbl = QLabel("Min rows:")
        mr_lbl.setStyleSheet(
            f"color: {G500}; font: 12px; background: transparent; border: none;"
        )
        mr_h.addWidget(mr_lbl)
        self._min_rows_spin = QSpinBox()
        self._min_rows_spin.setRange(1, 100)
        self._min_rows_spin.setValue(2)
        self._min_rows_spin.setFixedHeight(28)
        self._min_rows_spin.setStyleSheet(
            f"border: 1px solid {G200}; border-radius: 4px; font: 12px;"
            f" color: {G900}; background: {WHITE};"
        )
        mr_h.addWidget(self._min_rows_spin)
        mr_h.addStretch()
        lay.addWidget(min_row_w)
        lay.addSpacing(4)

        min_col_w = QWidget()
        min_col_w.setStyleSheet("background: transparent;")
        mc_h = QHBoxLayout(min_col_w)
        mc_h.setContentsMargins(0, 0, 0, 0)
        mc_h.setSpacing(8)
        mc_lbl = QLabel("Min columns:")
        mc_lbl.setStyleSheet(
            f"color: {G500}; font: 12px; background: transparent; border: none;"
        )
        mc_h.addWidget(mc_lbl)
        self._min_cols_spin = QSpinBox()
        self._min_cols_spin.setRange(1, 100)
        self._min_cols_spin.setValue(2)
        self._min_cols_spin.setFixedHeight(28)
        self._min_cols_spin.setStyleSheet(
            f"border: 1px solid {G200}; border-radius: 4px; font: 12px;"
            f" color: {G900}; background: {WHITE};"
        )
        mc_h.addWidget(self._min_cols_spin)
        mc_h.addStretch()
        lay.addWidget(min_col_w)
        lay.addSpacing(4)

        self._skip_image_chk = QCheckBox("Skip image-only pages")
        self._skip_image_chk.setChecked(True)
        self._skip_image_chk.setStyleSheet(
            f"color: {G700}; font: 12px; background: transparent;"
        )
        lay.addWidget(self._skip_image_chk)
        lay.addSpacing(16)

        # Output options
        self._section(lay, "OUTPUT OPTIONS")
        lay.addSpacing(8)

        sheet_row = QWidget()
        sheet_row.setStyleSheet("background: transparent;")
        sr_h = QHBoxLayout(sheet_row)
        sr_h.setContentsMargins(0, 0, 0, 0)
        sr_h.setSpacing(8)
        sr_lbl = QLabel("Sheet per:")
        sr_lbl.setStyleSheet(
            f"color: {G500}; font: 12px; background: transparent; border: none;"
        )
        sr_h.addWidget(sr_lbl)
        self._sheet_mode_combo = QComboBox()
        self._sheet_mode_combo.addItems(["Table", "Page"])
        self._sheet_mode_combo.setFixedHeight(28)
        self._sheet_mode_combo.setStyleSheet(self._combo_style())
        sr_h.addWidget(self._sheet_mode_combo, 1)
        lay.addWidget(sheet_row)
        lay.addSpacing(6)

        self._bold_header_chk = QCheckBox("Bold header row")
        self._bold_header_chk.setChecked(True)
        self._bold_header_chk.setStyleSheet(
            f"color: {G700}; font: 12px; background: transparent;"
        )
        lay.addWidget(self._bold_header_chk)
        lay.addSpacing(4)

        self._autofit_chk = QCheckBox("Auto-fit column widths")
        self._autofit_chk.setChecked(True)
        self._autofit_chk.setStyleSheet(
            f"color: {G700}; font: 12px; background: transparent;"
        )
        lay.addWidget(self._autofit_chk)

        lay.addStretch()
        outer_v.addWidget(scroll, 1)

        # Bottom action bar
        bottom = QWidget()
        bottom.setStyleSheet(f"background: {WHITE}; border-top: 1px solid {G200};")
        bot_lay = QVBoxLayout(bottom)
        bot_lay.setContentsMargins(24, 16, 24, 20)
        bot_lay.setSpacing(8)

        out_lbl = QLabel("OUTPUT FILE")
        out_lbl.setStyleSheet(
            f"color: {G500}; font: bold 11px; letter-spacing: 1.2px;"
            " background: transparent; border: none;"
        )
        bot_lay.addWidget(out_lbl)

        out_row = QHBoxLayout()
        out_row.setSpacing(6)
        self._out_entry = QLineEdit()
        self._out_entry.setPlaceholderText("output.xlsx")
        self._out_entry.setFixedHeight(34)
        self._out_entry.setStyleSheet(
            f"border: 1px solid {G200}; border-radius: 6px; padding: 0 10px;"
            f" font: 13px; color: {G900}; background: {WHITE};"
        )
        out_row.addWidget(self._out_entry, 1)
        out_browse_btn = QPushButton("Browse")
        out_browse_btn.setFixedSize(72, 34)
        out_browse_btn.setStyleSheet(
            f"QPushButton {{background: transparent; color: {G700};"
            f" border: 1px solid {G300}; border-radius: 6px; font: 13px;}}"
            f" QPushButton:hover {{background: {G200};}}"
        )
        out_browse_btn.clicked.connect(self._browse_output)
        out_row.addWidget(out_browse_btn)
        bot_lay.addLayout(out_row)

        self._status_lbl = QLabel("")
        self._status_lbl.setStyleSheet(
            f"color: {G500}; font: 12px; border: none; background: transparent;"
        )
        bot_lay.addWidget(self._status_lbl)

        self._progress = QProgressBar()
        self._progress.setFixedHeight(6)
        self._progress.setTextVisible(False)
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._progress.setStyleSheet(
            f"QProgressBar {{ background: {G200}; border-radius: 3px; border: none; }}"
            f"QProgressBar::chunk {{ background: {GREEN}; border-radius: 3px; }}"
        )
        self._progress.hide()
        bot_lay.addWidget(self._progress)

        self._extract_btn = QPushButton("Extract to Excel")
        self._extract_btn.setFixedHeight(42)
        self._extract_btn.setEnabled(False)
        self._extract_btn.setStyleSheet(
            f"QPushButton {{background: {GREEN}; color: white; border-radius: 6px;"
            f" font: bold 14px;}} QPushButton:hover {{background: {GREEN_HOVER};}}"
            f" QPushButton:disabled {{background: {G300}; color: {G500};}}"
        )
        self._extract_btn.clicked.connect(self._run_extraction)
        bot_lay.addWidget(self._extract_btn)

        outer_v.addWidget(bottom)

    def _build_right(self, parent_h: QHBoxLayout):
        right_container = QWidget()
        right_container.setStyleSheet(f"background: {WHITE};")
        right_v = QVBoxLayout(right_container)
        right_v.setContentsMargins(0, 0, 0, 0)
        right_v.setSpacing(0)
        parent_h.addWidget(right_container, 1)

        self._right_stack = QStackedWidget()
        right_v.addWidget(self._right_stack, 1)

        canvas_container = QWidget()
        canvas_container.setStyleSheet(f"background: {THUMB_BG};")
        cc_v = QVBoxLayout(canvas_container)
        cc_v.setContentsMargins(0, 0, 0, 0)
        cc_v.setSpacing(0)
        self._canvas = _PreviewCanvas(self)
        cc_v.addWidget(self._canvas)
        self._right_stack.addWidget(canvas_container)

        nav = QFrame()
        nav.setFixedHeight(44)
        nav.setStyleSheet(f"background: {G100}; border-top: 1px solid {G200};")
        nav_h = QHBoxLayout(nav)
        nav_h.setContentsMargins(10, 7, 10, 7)
        nav_h.setSpacing(4)

        prev_btn = QPushButton()
        prev_btn.setIcon(QIcon(svg_pixmap("chevron-left", G700, 14)))
        prev_btn.setIconSize(QSize(14, 14))
        prev_btn.setFixedSize(34, 30)
        prev_btn.setStyleSheet(
            f"QPushButton {{background: transparent; border-radius: 4px;}}"
            f" QPushButton:hover {{background: {G200};}}"
        )
        prev_btn.clicked.connect(self._prev_page)
        nav_h.addWidget(prev_btn)

        next_btn = QPushButton()
        next_btn.setIcon(QIcon(svg_pixmap("chevron-right", G700, 14)))
        next_btn.setIconSize(QSize(14, 14))
        next_btn.setFixedSize(34, 30)
        next_btn.setStyleSheet(
            f"QPushButton {{background: transparent; border-radius: 4px;}}"
            f" QPushButton:hover {{background: {G200};}}"
        )
        next_btn.clicked.connect(self._next_page)
        nav_h.addWidget(next_btn)

        self._page_lbl = QLabel("No file loaded")
        self._page_lbl.setStyleSheet(f"color: {G500}; font: 12px 'Segoe UI';")
        nav_h.addWidget(self._page_lbl)
        nav_h.addStretch()

        right_v.addWidget(nav)

    def _build_thumb_strip(self):
        strip_v = QVBoxLayout(self._strip_container)
        strip_v.setContentsMargins(0, 0, 0, 0)
        strip_v.setSpacing(0)

        self._thumb_scroll = QScrollArea()
        self._thumb_scroll.setWidgetResizable(True)
        self._thumb_scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded
        )
        self._thumb_scroll.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        self._thumb_scroll.setStyleSheet(f"border: none; background: {G100};")
        strip_v.addWidget(self._thumb_scroll)

        self._thumb_inner = QWidget()
        self._thumb_inner.setStyleSheet(f"background: {G100};")
        self._thumb_h_lay = QHBoxLayout(self._thumb_inner)
        self._thumb_h_lay.setContentsMargins(4, 8, 4, 8)
        self._thumb_h_lay.setSpacing(4)
        self._thumb_h_lay.addStretch()
        self._thumb_scroll.setWidget(self._thumb_inner)

        _WheelToHScroll(self._thumb_scroll)

    def _combo_style(self) -> str:
        return (
            f"QComboBox {{background: {WHITE}; color: {G700};"
            f" border: 1px solid {G200}; border-radius: 4px;"
            f" font: 12px; padding: 0 6px;}}"
            f" QComboBox::drop-down {{border: none;}}"
            f" QComboBox QAbstractItemView {{background: {WHITE}; color: {G700};"
            f" selection-background-color: {G100};}}"
        )

    def _section(self, parent_lay: QVBoxLayout, title: str):
        lbl = QLabel(title)
        lbl.setStyleSheet(
            f"color: {G500}; font: bold 11px; letter-spacing: 1.2px;"
            " background: transparent; border: none;"
        )
        parent_lay.addWidget(lbl)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open PDF", "", "PDF files (*.pdf)")
        if path:
            self._open_pdf(path)

    def _browse_output(self):
        current = self._out_entry.text().strip()
        start = str(Path(self.pdf_path).parent) if self.pdf_path else ""
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(start, current) if current and start else current,
            "Excel Files (*.xlsx)",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self._out_entry.setText(path)

    def _open_pdf(self, path: str):
        if self._doc:
            try:
                self._doc.close()
            except RuntimeError:
                pass
        if self._pldoc:
            try:
                self._pldoc.close()
            except RuntimeError:
                pass
        self._doc = None
        self._pldoc = None
        self._password = ""

        try:
            doc = fitz.open(path)
        except Exception as e:
            logger.exception("could not open pdf")
            QMessageBox.critical(self, "Cannot Open PDF", f"Failed to open file:\n{e}")
            return

        if doc.needs_pass:
            pw, ok = QInputDialog.getText(
                self,
                "Password Required",
                "This PDF is password-protected.\nEnter password:",
                QLineEdit.EchoMode.Password,
            )
            if not ok:
                doc.close()
                return
            if not doc.authenticate(pw):
                QMessageBox.critical(
                    self, "Wrong Password", "Incorrect password. Cannot open PDF."
                )
                doc.close()
                return
            self._password = pw

        if doc.page_count == 0:
            QMessageBox.critical(self, "Empty PDF", "This PDF contains no pages.")
            doc.close()
            return

        try:
            if self._password:
                self._pldoc = pdfplumber.open(path, password=self._password)
            else:
                self._pldoc = pdfplumber.open(path)
        except Exception as e:
            logger.exception("could not open pdf")
            QMessageBox.critical(
                self, "pdfplumber Error", f"Could not open PDF with pdfplumber:\n{e}"
            )
            doc.close()
            return

        self._doc = doc
        self.pdf_path = path
        self._total_pages = doc.page_count
        self._current_page = 0

        stem = Path(path).stem
        self._out_entry.setText(f"{stem}.xlsx")
        self._file_entry.setText(os.path.basename(path))
        self._extract_btn.setEnabled(True)
        self._status_lbl.setText(
            f"{doc.page_count} page{'s' if doc.page_count != 1 else ''} loaded."
        )

        self._build_thumbnails()
        self._render_page_canvas()

    def _page_is_image_only(self, page_idx: int) -> bool:
        if not self._doc:
            return False
        raw = self._doc[page_idx].get_text()
        text = raw.strip() if isinstance(raw, str) else ""
        return len(text) < 5

    def _build_thumbnails(self):
        while self._thumb_h_lay.count() > 1:
            item = self._thumb_h_lay.takeAt(0)
            if item is not None:
                w = item.widget()
                if w is not None:
                    w.deleteLater()

        self._thumb_pixmaps = []
        self._highlighted_thumb_frame = None
        self._thumb_render_next = 0

        for i in range(self._total_pages):
            frame = QWidget()
            frame.setStyleSheet(f"background: {G100};")
            frame_v = QVBoxLayout(frame)
            frame_v.setContentsMargins(0, 0, 0, 0)
            frame_v.setSpacing(2)
            frame.setFixedWidth(self.THUMB_W + 8)

            thumb_lbl = QLabel()
            thumb_lbl.setFixedSize(self.THUMB_W, 110)
            thumb_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            thumb_lbl.setStyleSheet(
                f"background: {THUMB_BG}; border: 1px solid {G300};"
            )
            frame_v.addWidget(thumb_lbl, 0, Qt.AlignmentFlag.AlignHCenter)

            num_lbl = QLabel(str(i + 1))
            num_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            num_lbl.setStyleSheet(f"color: {G500}; font: 9px 'Segoe UI';")
            frame_v.addWidget(num_lbl)

            idx_capture = i
            thumb_lbl.mousePressEvent = lambda e, idx=idx_capture: self._go_to_page(idx)
            num_lbl.mousePressEvent = lambda e, idx=idx_capture: self._go_to_page(idx)
            frame.mousePressEvent = lambda e, idx=idx_capture: self._go_to_page(idx)

            self._thumb_h_lay.insertWidget(self._thumb_h_lay.count() - 1, frame)
            self._thumb_pixmaps.append((None, frame, thumb_lbl))

        self._render_thumb_batch()

    def _render_thumb_batch(self, batch: int = 8):
        if not self._doc:
            return
        start = self._thumb_render_next
        end = min(start + batch, self._total_pages)
        for i in range(start, end):
            pm_old, frame, lbl = self._thumb_pixmaps[i]
            if pm_old is not None:
                continue
            try:
                pm = _render_thumb_qpixmap(self._doc, i, self.THUMB_W)
            except RuntimeError:
                continue
            self._thumb_pixmaps[i] = (pm, frame, lbl)
            lbl.setPixmap(
                pm.scaled(
                    self.THUMB_W,
                    110,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
            )
        self._thumb_render_next = end
        if end < self._total_pages:
            QTimer.singleShot(0, self._render_thumb_batch)
        else:
            self._highlight_thumb(self._current_page)

    def _highlight_thumb(self, idx: int):
        if self._highlighted_thumb_frame is not None:
            try:
                for entry in self._thumb_pixmaps:
                    if entry[1] is self._highlighted_thumb_frame:
                        entry[2].setStyleSheet(
                            f"background: {THUMB_BG}; border: 1px solid {G300};"
                        )
                        break
            except RuntimeError:
                pass
        if 0 <= idx < len(self._thumb_pixmaps):
            _, frame, lbl = self._thumb_pixmaps[idx]
            lbl.setStyleSheet(f"background: {THUMB_BG}; border: 2px solid {BLUE};")
            self._highlighted_thumb_frame = frame
        else:
            self._highlighted_thumb_frame = None

    def _render_page_canvas(self):
        if not self._doc:
            return
        cw = max(self._canvas.width(), 100)
        try:
            pm, scale = _render_page_qpixmap(self._doc, self._current_page, cw - 20)
        except RuntimeError:
            return
        self._page_pixmap = pm
        self._page_scale = scale
        iw = pm.width()
        self._page_ox = (cw - iw) // 2
        self._page_oy = 10
        self._page_lbl.setText(f"Page {self._current_page + 1} / {self._total_pages}")
        self._draw_table_outlines()
        self._highlight_thumb(self._current_page)
        self._canvas.update()

    def _draw_table_outlines(self):
        self._canvas_table_rects = []
        if not self._pldoc:
            return
        try:
            pl_page = self._pldoc.pages[self._current_page]
            tables = pl_page.find_tables(self._build_table_settings())
        except (ValueError, RuntimeError):
            return
        self._table_bboxes = []
        for table in tables:
            x0, top, x1, bottom = table.bbox
            self._table_bboxes.append((x0, top, x1, bottom))
            cx0 = self._page_ox + x0 * self._page_scale
            cy0 = self._page_oy + top * self._page_scale
            cx1 = self._page_ox + x1 * self._page_scale
            cy1 = self._page_oy + bottom * self._page_scale
            self._canvas_table_rects.append((cx0, cy0, cx1, cy1))

    def _go_to_page(self, idx: int):
        if not self._doc:
            return
        idx = max(0, min(idx, self._total_pages - 1))
        self._current_page = idx
        self._render_page_canvas()

    def _prev_page(self):
        self._go_to_page(self._current_page - 1)

    def _next_page(self):
        self._go_to_page(self._current_page + 1)

    def _build_table_settings(self) -> dict:
        method_text = self._detect_combo.currentText()
        if "Lattice" in method_text:
            v_strat = "lines"
            h_strat = "lines"
        elif "Stream" in method_text:
            v_strat = "text"
            h_strat = "text"
        else:
            v_strat = "lines_strict"
            h_strat = "lines_strict"
        return {
            "vertical_strategy": v_strat,
            "horizontal_strategy": h_strat,
            "intersection_y_tolerance": 3,
            "intersection_x_tolerance": 3,
            "snap_y_tolerance": 3,
            "snap_x_tolerance": 3,
            "edge_min_length": 3,
            "min_words_vertical": 1,
            "min_words_horizontal": 1,
            "keep_blank_chars": False,
            "text_tolerance": 3,
            "text_x_tolerance": 3,
            "text_y_tolerance": 3,
            "explicit_vertical_lines": [],
            "explicit_horizontal_lines": [],
        }

    @staticmethod
    def _parse_page_range(spec: str, total: int) -> list:
        spec = spec.strip().lower()
        if spec in ("", "all"):
            return list(range(total))
        pages: set = set()
        for part in spec.split(","):
            part = part.strip()
            if "-" in part:
                lo, _, hi = part.partition("-")
                lo_i = int(lo.strip()) - 1
                hi_i = int(hi.strip()) - 1
                if lo_i < 0 or hi_i >= total or lo_i > hi_i:
                    raise ValueError(
                        f"Page range '{part}' is out of bounds "
                        f"(document has {total} pages)."
                    )
                pages.update(range(lo_i, hi_i + 1))
            else:
                idx = int(part) - 1
                if idx < 0 or idx >= total:
                    raise ValueError(
                        f"Page number {part} is out of bounds "
                        f"(document has {total} pages)."
                    )
                pages.add(idx)
        return sorted(pages)

    def _run_extraction(self):
        if not self._doc or not self._pldoc:
            QMessageBox.warning(self, "No File", "Please open a PDF file first.")
            return

        out_name = self._out_entry.text().strip()
        if not out_name:
            QMessageBox.warning(
                self, "No Output File", "Please specify an output file name."
            )
            return
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"

        if not os.path.isabs(out_name) and not os.path.dirname(out_name):
            out_path = os.path.join(str(Path(self.pdf_path).parent), out_name)
        else:
            out_path = out_name

        range_text = self._range_entry.text().strip()
        try:
            pages = self._parse_page_range(range_text, self._total_pages)
        except ValueError as e:
            QMessageBox.critical(self, "Invalid Page Range", str(e))
            return

        if not pages:
            QMessageBox.warning(self, "Empty Selection", "No pages to process.")
            return

        min_rows = self._min_rows_spin.value()
        min_cols = self._min_cols_spin.value()
        skip_image = self._skip_image_chk.isChecked()
        sheet_mode = self._sheet_mode_combo.currentText()
        bold_header = self._bold_header_chk.isChecked()
        auto_fit = self._autofit_chk.isChecked()
        settings = self._build_table_settings()

        self._extract_btn.setEnabled(False)
        self._extract_btn.setText("Extracting…")
        self._progress.setValue(0)
        self._progress.show()
        self._status_lbl.setText("Starting…")

        self._worker = _ExcelExtractionWorker(
            pdf_path=self.pdf_path,
            password=self._password,
            pages=pages,
            out_path=out_path,
            settings=settings,
            min_rows=min_rows,
            min_cols=min_cols,
            skip_image=skip_image,
            sheet_mode=sheet_mode,
            bold_header=bold_header,
            auto_fit=auto_fit,
        )
        self._worker.progress.connect(self._on_extraction_progress)
        self._worker.finished.connect(self._on_extraction_done)
        self._worker.failed.connect(self._on_extraction_failed)
        self._worker.start()

    def _on_extraction_progress(self, pct: int, text: str):
        self._progress.setValue(pct)
        self._status_lbl.setText(text)

    def _on_extraction_done(self, report_text: str, out_dir: str):
        self._extract_btn.setEnabled(True)
        self._extract_btn.setText("Extract to Excel")
        self._progress.setValue(100)
        lines = report_text.split("\n")
        n_t = sum(1 for ln in lines if ln.startswith("Table "))
        self._status_lbl.setText(
            f"Done. {n_t} table{'s' if n_t != 1 else ''} extracted."
        )
        self._status_lbl.setStyleSheet(
            f"color: {EMERALD}; font: 12px; border: none; background: transparent;"
        )
        self._show_report(report_text, out_dir)

    def _on_extraction_failed(self, msg: str):
        self._extract_btn.setEnabled(True)
        self._extract_btn.setText("Extract to Excel")
        self._progress.hide()
        QMessageBox.critical(self, "Extraction Failed", msg)

    def _show_report(self, text: str, output_dir: str):
        if self._report_widget:
            self._report_widget.deleteLater()

        report = QWidget()
        report.setStyleSheet(f"background: {WHITE};")
        self._report_widget = report

        v = QVBoxLayout(report)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"border: none; background: {WHITE};")
        inner = QWidget()
        inner.setStyleSheet(f"background: {WHITE};")
        inner_lay = QVBoxLayout(inner)
        inner_lay.setContentsMargins(20, 16, 20, 16)
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setStyleSheet(
            f"color: {G700}; font: 12px 'Courier New'; background: transparent;"
        )
        inner_lay.addWidget(lbl)
        inner_lay.addStretch()
        scroll.setWidget(inner)
        v.addWidget(scroll, 1)

        btn_bar = QFrame()
        btn_bar.setFixedHeight(50)
        btn_bar.setStyleSheet(f"background: {G100}; border-top: 1px solid {G200};")
        btn_bar_lay = QHBoxLayout(btn_bar)
        btn_bar_lay.setContentsMargins(16, 8, 16, 8)

        open_btn = QPushButton("Open Output Folder")
        open_btn.setFixedSize(180, 34)
        open_btn.setStyleSheet(
            f"QPushButton {{background: {BLUE}; color: white; border-radius: 6px;"
            f" font: 13px;}} QPushButton:hover {{background: {BLUE_HOVER};}}"
        )
        open_btn.clicked.connect(lambda: self._open_folder(output_dir))
        btn_bar_lay.addWidget(open_btn)

        back_btn = QPushButton("\u2190 Back to Preview")
        back_btn.setFixedSize(160, 34)
        back_btn.setStyleSheet(
            f"QPushButton {{background: transparent; color: {G700};"
            f" border: 1px solid {G300}; border-radius: 6px; font: 13px;}}"
            f" QPushButton:hover {{background: {G200};}}"
        )
        back_btn.clicked.connect(self._back_to_preview)
        btn_bar_lay.addWidget(back_btn)
        btn_bar_lay.addStretch()
        v.addWidget(btn_bar)

        self._right_stack.addWidget(report)
        self._right_stack.setCurrentWidget(report)

    def _back_to_preview(self):
        self._right_stack.setCurrentIndex(0)
        if self._report_widget:
            self._report_widget.deleteLater()
            self._report_widget = None
        if self._doc:
            self._render_page_canvas()

    def _open_folder(self, path: str):
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", path])
            elif sys.platform == "win32":
                os.startfile(path)
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            logger.exception("could not open folder")
            QMessageBox.critical(self, "Error", f"Could not open folder:\n{e}")

    def cleanup(self):
        if self._thumb_timer is not None:
            try:
                self._thumb_timer.stop()
            except RuntimeError:
                pass
            self._thumb_timer = None
        if self._doc is not None:
            try:
                self._doc.close()
            except RuntimeError:
                pass
        if self._pldoc is not None:
            try:
                self._pldoc.close()
            except RuntimeError:
                pass
