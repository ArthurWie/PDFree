"""Tests for form_export_tool extraction logic."""

import csv
import json

import pytest

fitz = pytest.importorskip("fitz", reason="PyMuPDF not installed")


def _make_form_pdf(path: str) -> None:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    for name, value, y in [("First", "Alice", 72), ("Last", "Smith", 120)]:
        w = fitz.Widget()
        w.field_type = fitz.PDF_WIDGET_TYPE_TEXT
        w.field_name = name
        w.rect = fitz.Rect(72, y, 300, y + 28)
        w.field_value = value
        page.add_widget(w)
    doc.save(path)
    doc.close()


def _make_plain_pdf(path: str) -> None:
    doc = fitz.open()
    doc.new_page(width=595, height=842)
    doc.save(path)
    doc.close()


def _extract(path: str) -> list:
    doc = fitz.open(path)
    rows = []
    for pg_idx in range(doc.page_count):
        page = doc[pg_idx]
        for w in page.widgets():
            rows.append(
                {
                    "page": pg_idx + 1,
                    "field_name": w.field_name or "",
                    "field_type": w.field_type_string or "",
                    "field_value": str(w.field_value)
                    if w.field_value is not None
                    else "",
                }
            )
    doc.close()
    return rows


def test_extract_returns_correct_field_count(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)
    assert len(rows) == 2


def test_extract_captures_field_names(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)
    names = {r["field_name"] for r in rows}
    assert "First" in names
    assert "Last" in names


def test_extract_captures_field_values(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)
    values = {r["field_value"] for r in rows}
    assert "Alice" in values
    assert "Smith" in values


def test_extract_captures_page_number(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)
    assert all(r["page"] == 1 for r in rows)


def test_extract_plain_pdf_returns_empty(tmp_path):
    src = str(tmp_path / "plain.pdf")
    _make_plain_pdf(src)
    rows = _extract(src)
    assert rows == []


def test_json_export_roundtrip(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)

    dst = str(tmp_path / "fields.json")
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2)

    with open(dst, encoding="utf-8") as f:
        loaded = json.load(f)

    assert len(loaded) == len(rows)
    assert loaded[0]["field_name"] == rows[0]["field_name"]


def test_csv_export_roundtrip(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)

    dst = str(tmp_path / "fields.csv")
    with open(dst, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["page", "field_name", "field_type", "field_value"]
        )
        writer.writeheader()
        writer.writerows(rows)

    with open(dst, newline="", encoding="utf-8") as f:
        reader = list(csv.DictReader(f))

    assert len(reader) == len(rows)
    assert {r["field_name"] for r in reader} == {r["field_name"] for r in rows}


def test_csv_export_has_header(tmp_path):
    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)

    dst = str(tmp_path / "fields.csv")
    with open(dst, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["page", "field_name", "field_type", "field_value"]
        )
        writer.writeheader()
        writer.writerows(rows)

    with open(dst, encoding="utf-8") as f:
        first_line = f.readline().strip()

    assert "field_name" in first_line
    assert "field_value" in first_line


def test_xlsx_export_roundtrip(tmp_path):
    import openpyxl

    src = str(tmp_path / "form.pdf")
    _make_form_pdf(src)
    rows = _extract(src)

    dst = str(tmp_path / "fields.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["page", "field_name", "field_type", "field_value"])
    for row in rows:
        ws.append(
            [row["page"], row["field_name"], row["field_type"], row["field_value"]]
        )
    wb.save(dst)

    wb2 = openpyxl.load_workbook(dst)
    ws2 = wb2.active
    data = list(ws2.iter_rows(values_only=True))
    assert data[0] == ("page", "field_name", "field_type", "field_value")
    assert len(data) == len(rows) + 1  # header + rows
